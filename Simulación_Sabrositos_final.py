#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Simulación de inventario 'Sabrositos' con R réplicas para determinar costo promedio diario para
distintas cantidades de stock a mantener en el inventario.
----------------------------------------------------------------------
• Genera N números Uniforme[0,1) (LCG) y les aplica 4 pruebas
• Un número → #personas; luego k números → demanda individual
• R réplicas de 2 meses (46 días hábiles) cada una
• Exporta a Excel: hoja 'Simulación' + hoja 'Pruebas' para cada variable de control (stock a mantener de Sabrositos al reponer
con proveedor)+ hoja 'Conclusiones' con los intervalos de confianza.
"""

import math
from typing import Dict, List, Tuple
import numpy as np
import pandas as pd
from scipy import stats

# ---------------------------------------------------------------------------
# 1. Generador congruencial mixto
# ---------------------------------------------------------------------------
def lcg(seed: int, n: int,
        a: int = 1_103_515_245, c: int = 12_345, m: int = 2**31) -> np.ndarray:
    x, nums = seed, np.empty(n, dtype=float)
    for i in range(n):
        x = (a * x + c) % m
        nums[i] = x / m
    return nums


# ---------------------------------------------------------------------------
# 2. Pruebas de aleatoriedad
# ---------------------------------------------------------------------------
def chi_square_test(nums: np.ndarray, k: int = 10, alpha: float = .05):
    n = len(nums)
    freq = np.histogram(nums, bins=k, range=(0, 1))[0]
    exp = n / k
    chi = ((freq - exp) ** 2 / exp).sum()
    crit = stats.chi2.ppf(1 - alpha, k - 1)
    p = 1 - stats.chi2.cdf(chi, k - 1)
    if chi < crit:
        rechazoHo="NO"
    else:
        rechazoHo="SÍ"
    return chi, crit, p, rechazoHo


def mean_test(nums: np.ndarray, alpha: float = .05):
    n = len(nums)
    z = (nums.mean() - .5) * math.sqrt(12 * n)
    zcrit = stats.norm.ppf(1 - alpha / 2)
    if abs(z) < zcrit:
        rechazoHo="NO"
    else:
        rechazoHo="SÍ"
    return z, zcrit, rechazoHo


def variance_test(nums: np.ndarray, alpha: float = .05):
    n = len(nums)
    chi = (n - 1) * nums.var(ddof=1) / (1 / 12)
    chi_inf = stats.chi2.ppf(alpha / 2, n - 1)
    chi_sup = stats.chi2.ppf(1 - alpha / 2, n - 1)
    if (chi_inf < chi < chi_sup):
        rechazoHo="NO"
    else:
        rechazoHo="SÍ"
    return chi, chi_inf, chi_sup,rechazoHo 


def ks_uniformity_test(nums: np.ndarray, alpha: float = .05):
    n = len(nums)
    d, p = stats.kstest(nums, 'uniform')
    dcrit = stats.kstwobign.ppf(1 - alpha) / math.sqrt(n)
    if (d < dcrit):
        rechazoHo="NO"
    else:
        rechazoHo="SÍ"
    return d, dcrit, p, rechazoHo

def run_all_tests(nums: np.ndarray, alpha: float = .05) -> Dict[str, Dict]:
    chi, crit, p, ok = chi_square_test(nums, alpha=alpha)
    z, zcrit, okm = mean_test(nums, alpha=alpha)
    chiv, ci, cs, okv = variance_test(nums, alpha=alpha)
    d, dcrit, pks, okks = ks_uniformity_test(nums, alpha=alpha)
    return {
        'Chi-cuadrado': {'Estadístico': chi, 'Valor crítico': crit,
                         'p-value': p, '¿Rechazo H₀?': ok},
        'Media':        {'Estadístico (z)': z, 'Valor crítico ±': zcrit,
                         'p-value': 2*(1 - stats.norm.cdf(abs(z))), '¿Rechazo H₀?': okm},
        'Varianza':     {'Estadístico (χ²)': chiv,
                         'Rango crítico': f"[{ci:.3f}, {cs:.3f}]", '¿Rechazo H₀?': okv},
        'Kolmogórov-Smirnov': {'Estadístico D': d, 'Valor crítico': dcrit,
                               'p-value': pks, '¿Rechazo H₀?': okks}
    }

# ---------------------------------------------------------------------------
# 3. Mapeos de distribución
# ---------------------------------------------------------------------------
def personas_por_uniforme(u: float) -> int:
    if u < 0.21871189:
        return 0
    elif u < 0.55115396:
        return 1
    elif u < 0.80380993:
        return 2
    elif u < 0.93182229:
        return 3
    elif u < 0.98046698:
        return 4
    else:
        return 5


def demanda_por_uniforme(u: float) -> int:
    if u < 0.11428571:
        return 400
    elif u < 0.42857143:
        return 800
    elif u < 0.94285714:
        return 1000
    else:
        return 1500


# ---------------------------------------------------------------------------
# 4. Parámetros de ALMACENAMIENTO
# ---------------------------------------------------------------------------
ALQUILER_MENSUAL    = 300_000        # $/mes del local completo
M2_LOCAL            = 50            # m² totales del local
BOLSA_M2            = 0.5            # cada bolsa ocupa 0,5 m²
CAPACIDAD_GR        = 20000         #Capacidad de la bolsa que ocupa los 0,5 m²

COSTO_ALM_xGramo       = ALQUILER_MENSUAL*((BOLSA_M2)/M2_LOCAL) / CAPACIDAD_GR



# ---------------------------------------------------------------------------
# 5. Función para llegar al intervalos de confianza correspondiente a cada valor de la variable de control
# ---------------------------------------------------------------------------

def llegar_a_intervalos(data:list, alpha: float = 0.05):
    réplicas = len(data)
    media=sum(data)/réplicas
    varianza=np.var(data) #obtenemos varianza 
    desvío_estándar=math.sqrt(varianza)
    
    # Lo necesario para demostrar si sigue una distribución normal
    dcrit = stats.kstwobign.ppf(1 - alpha) / math.sqrt(réplicas)
    d, p = stats.kstest(data, 'norm', args=(media, desvío_estándar))
    demuestra_normalidad=d<dcrit

    if (demuestra_normalidad):
        rechazoHo="NO"
        error_estándar = desvío_estándar/ math.sqrt(réplicas)
        # Valor crítico de la distribución normal, lo ocupamos en vez de la fórmula t de Student porque la muestra es mayor a 30( ya que hicimos 50 réplicas)
        z_critico = stats.norm.ppf(1 - alpha/2)
        # Calcular límites del intervalo
        limite_inferior = media - error_estándar * z_critico
        limite_superior = media + error_estándar * z_critico
    else:
        rechazoHo="SÍ"
        limite_inferior= media - desvío_estándar/math.sqrt(réplicas*alpha)
        limite_superior= media + desvío_estándar/math.sqrt(réplicas*alpha)

    # Retorna: estadístico KS, valor crítico, valor p, si se rechaza H₀ y el intervalo de confianza para este valor de la variable de control.
    return {
        '¿Rechazo H₀ de normalidad?': rechazoHo,'Intervalo de confianza(LI)':limite_inferior,'Intervalo de confianza(LS)':limite_superior 
        }


# ---------------------------------------------------------------------------
# 6. Función para las réplicas correspondientes a un valor de la variable de control
# ---------------------------------------------------------------------------
columnas = [
        'Réplica',
        'Día',
        'u_personas',
        'Personas',
        'u_demandas',
        'Demanda(gramos)',
        'Inventario inicial',
        'Inventario final',
        'Costo de Compra',
        'Costo Faltante',
        'Costo Almacenamiento',
        'Costo Total',
        'Ventas'
    ]

def simular_replicas(u: np.ndarray, replicas: int,
                     dias_por_replica: int,
                     inventario_normal: int,
                     costo_normal: float,
                     costoxgramo: float) -> pd.DataFrame:

    filas: List[dict] = []
    idx = 0
    max_necesarios = replicas * dias_por_replica * 5
    
    if len(u) < max_necesarios:
        raise ValueError(f"Se requieren al menos {max_necesarios} números aleatorios.")
    
    costos_prom_stock_a_probar=[] #Se usará para determinar los intervalos
    for r in range(1, replicas + 1):
        inv = inventario_normal
        costototalacumulado = 0 # está para calcular el promedio de costo total correspondiente a la réplica
        if r !=1:
            filas.append({col: col for col in columnas})
        for d in range(1, dias_por_replica + 1):

            # reposición cada 12 días (incl. día 1)
            if d == 1 or d % 12 == 1:
                pct = (inventario_normal - inv) / inventario_normal
                costo_compra = pct * costo_normal
                inv = inventario_normal
            else:
                costo_compra = 0.0

            # número de personas
            u_personas = float(u[idx]); idx += 1
            personas = personas_por_uniforme(u_personas)

            # demandas individuales
            demanda_total = 0
            u_dem_list = []
            for _ in range(personas):
                u_dem = float(u[idx]); idx += 1
                u_dem_list.append(u_dem)
                demanda_total += demanda_por_uniforme(u_dem)
            u_dem_str = ";".join(f"{v:.6f}" for v in u_dem_list)

            # inventario y faltante
            if inv >= demanda_total:
                inv_fin = inv - demanda_total
                costo_faltante = 0.0
                ventas= demanda_total
            else:
                ventas = inv 
                inv_fin = 0
                costo_faltante = (demanda_total - inv) * (2500/1000) #El costo de faltante lo consideraremos como lo que hubiese ingresado si tuviera stock
            # Costo de almacenacenamiento en el que incurro en el día, sacado a partir del inventario promedio en el día.
            costo_almacen = COSTO_ALM_xGramo*((inv+inv_fin)/2) 

            # costo total diario
            costo_total = costo_compra + costo_faltante + costo_almacen

            costototalacumulado+=costo_total
            
            filas.append({
                'Réplica': r,
                'Día': d,
                'u_personas': u_personas,
                'Personas': personas,
                'u_demandas': u_dem_str,
                'Demanda(gramos)': demanda_total,
                'Inventario inicial': inv,
                'Inventario final': inv_fin,
                'Costo de Compra': costo_compra,
                'Costo Faltante': costo_faltante, #owo 
                'Costo Almacenamiento': costo_almacen, #anadido owo 
                'Costo Total': costo_total, #owo al cuadrado
                'Ventas': ventas, #anadido uwu 
            })

            inv = inv_fin
        costopromedio=costototalacumulado/dias_por_replica
        
        filas.append({  #ignorar las keys, es solo para no complejizar la manera de que aparezca el costo promedio luego de cada réplica
            'Réplica':"costo promedio", 
            'Día':(costopromedio)
        })
        costos_prom_stock_a_probar.append(costopromedio) 
    return pd.DataFrame(filas), costos_prom_stock_a_probar


# ---------------------------------------------------------------------------
# 7. Exportar a Excel Simulación y resultados de pruebas a los números aleatorios
# ---------------------------------------------------------------------------
def ir_guardando_en_excel(creararchivo,
                     df_sim: pd.DataFrame,
                     resultados_pruebas: Dict[str, Dict],stock_a_probar,
                     filename: str) -> None:

    df_pruebas = (pd.DataFrame(resultados_pruebas)
                .T.reset_index()
                .rename(columns={'index': 'Prueba'}))

    if not creararchivo: 
        with pd.ExcelWriter(filename,engine='openpyxl',mode='a') as w:
            df_sim.to_excel(w, sheet_name=f"SimulaciónPara{stock_a_probar}", index=False)
            df_pruebas.to_excel(w, sheet_name=f"PruebasPara{stock_a_probar}",index=False)
            
    else:
        with pd.ExcelWriter(filename) as w:
            df_sim.to_excel(w, sheet_name=f"SimulaciónPara{stock_a_probar}", index=False)
            df_pruebas.to_excel(w, sheet_name=f"PruebasPara{stock_a_probar}",     index=False)

# ---------------------------------------------------------------------------
# 8. Una función para ajustar el ancho de las columnas basándome solo en los encabezados para no tardar mucho.
# ---------------------------------------------------------------------------
from openpyxl.utils import get_column_letter

def autofit_encabezados(worksheet, margin=2, scale=1.2):
    """
    Ajusta el ancho de las columnas basado solo en los encabezados
    - worksheet: Hoja de trabajo de OpenPyXL
    - margin: Margen adicional (default 2)
    - scale: Factor de escala (default 1.2)
    """
    for col_idx, column in enumerate(worksheet.iter_cols(), 1):
        if column[0].value:  # Solo si hay valor en el encabezado
            col_letter = get_column_letter(col_idx)
            ancho_encabezado = len(str(column[0].value))
            ancho_ajustado = (ancho_encabezado + margin) * scale
            worksheet.column_dimensions[col_letter].width = ancho_ajustado            
            
        


# ---------------------------------------------------------------------------
# 9. Ejecución principal
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    N  = int(input("¿Cuántos números pseudoaleatorios generar?(Por defecto 400.000)    ") or 400000) #(48 días)X(máximo 5 personas en un día)X(1600 réplicas)= máximo 384.000 números necesitamos, así que igual por las dudas le paso 400.000 XD
    R  = int(input("¿Cuántas réplicas de 2 meses desea?(Por defecto 1.600)         ") or 1600)

    # datos del problema
    inventario_conocido = 20_000 # Es decir que conocemos cuánto hay que pagar la bolsa de 20 kilos al proveedor
    costo_conocido = 38_323
    costoxgramo = costo_conocido / inventario_conocido
    
    nombrearchivo="SimulacionSabrositos.xlsx"
    creararchivo=True
    stocks_de_prueba_y_seed=[[12_000,456789],[9_000,654321],[5_000,234567]]
    para_intervalos={}
    for stock_a_probar,seed in stocks_de_prueba_y_seed:
        # generación de números y pruebas
        u = lcg(seed, N)
        print('...')
        res_pruebas = run_all_tests(u)
        # simulación
        costo_normal = costoxgramo * stock_a_probar
        df_sim,costos_prom_stock_a_probar = simular_replicas(u, R, 48,
                              stock_a_probar, costo_normal, costoxgramo)

         # exportar Simulación y pruebas a los números aleatorios
        ir_guardando_en_excel(creararchivo,df_sim,res_pruebas,stock_a_probar,nombrearchivo)
        para_intervalos[stock_a_probar]=llegar_a_intervalos(costos_prom_stock_a_probar)
        creararchivo=False
    
    #Terminar la exportación a Excel, añadiendo la hoja correspondiente a los intervalos de confianza.
    if not creararchivo:
        df_int=(pd.DataFrame(para_intervalos)
                    .T.reset_index()
                    .rename(columns={'index': 'Stock probado'}))
        with pd.ExcelWriter(nombrearchivo,engine='openpyxl',mode='a') as w:
            df_int.to_excel(w, sheet_name="Intervalos resultantes",index=False)
    
    # Ajusta solo los encabezados
    with pd.ExcelWriter(nombrearchivo, engine='openpyxl',mode='a') as w:
        for sheet_name, worksheet in w.sheets.items():
            autofit_encabezados(worksheet)  

    print("✓ Simulación terminada: 'SimulacionSabrositos.xlsx'")
